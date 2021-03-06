# -*- coding: utf-8 -*-
"""twitter

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1HzsJVKw6rXVQ7tP8J-tdL_NDi0oYjChl
"""

"""!pip3 install nltk
!pip3 install pandas
!pip3 install numpy

!pip3 install spacy
!pip3 install corpus
!pip3 install stem
!pip3 install snowball

!pip3 install wordnet
!pip3 install query
!pip3 install snscrape"""

#!/usr/bin/env python
# -*- coding: utf-8 -*-

import snscrape.modules.twitter as sntwitter
import nltk
#nltk.download('punkt')
#nltk.download('stopwords')
import string
from nltk import sent_tokenize, word_tokenize
from nltk.stem.snowball import SnowballStemmer
from nltk.stem.wordnet import WordNetLemmatizer
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import pandas as pd
import numpy as np
import re  
import spacy
import datetime as dt 
maxTweet = 1000
dosya = open('twet.txt','r+')#
for i, tweet in enumerate(sntwitter.TwitterSearchScraper('from:TechBBQ + since:2009-05-27 until:2021-01-8').get_items()):
  if i>maxTweet:
    break 
  yolla = tweet.content.strip('twet.txt')
  dosya.write(yolla+ "\n")  
print(i)
dosya.close()

import re
dosya = open('twet.txt','r').read()
text1 = dosya
metin = re.sub("[^a-zA-Z]", " ", text1)
metin = metin.lower()
bol = metin.split()
birlestir = " ".join(bol)
print(bol)

import nltk
#nltk.download('stopwords')
from nltk.corpus import stopwords

stopwords = open('english.txt', 'r').read()

birlestir1 = []
for word in bol:
  if not word in stopwords :
    birlestir1.append(word)
print(birlestir1)

from openpyxl import Workbook,load_workbook
wb = load_workbook("kelimeler.xlsx")
ws = wb.active
text = birlestir1
kayit = []
for satir in range(2,88):
  for word in text:
    if word == (ws.cell(satir,1).value):
      kayit.append(word)
      print(word)#word e append yaparsan listeye atarsın :)

print((kayit))

from nltk.stem.snowball import SnowballStemmer

s_stemmer = SnowballStemmer(language='english')
stem = []

for word in kayit:
    stem.append(s_stemmer.stem(word))


    

print(stem)
print(kayit)

"""!pip install seaborn"""

import matplotlib.pyplot as plt; plt.rcdefaults()
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

print(kayit)

### sonuçlar
kayit = pd.DataFrame(stem)
kayit = kayit[0].value_counts()

kayit = kayit[:20]
plt.figure(figsize=(10,5))
sns.barplot(kayit.values,kayit.index,alpha = 1)
plt.title('Top Words')
plt.ylabel("Word from Tweet",fontsize = 12)
plt.xlabel('Count of Words',fontsize = 12)
plt.show()
print(kayit)

"""!python -m spacy download en_core_web_lg"""

import spacy
from spacy import displacy
from collections import Counter
import en_core_web_sm
nlp = en_core_web_sm.load()

def show_ents(doc):
    if doc.ents:
        for ent in doc.ents:
            print(ent.text + ' - ' + ent.label_ + ' - ' + str(spacy.explain(ent.label_)))

str1 = " " 
top_org = str1.join(birlestir1)

top_org = nlp(top_org)

label = [(X.text, X.label_) for X in top_org.ents]
print(label)

df1 = pd.DataFrame(label, columns = ['Word','Entity'])

df2 = df1.where(df1['Entity'] == 'ORG')
 
df2 = df2['Word'].value_counts()
print(df2)

#################
top_org = df2[:10,]
plt.figure(figsize=(10,5))
sns.barplot(top_org.values, top_org.index, alpha=0.8)
plt.title('Top Organizations Mentioned')
plt.ylabel('Word from Tweet', fontsize=12)
plt.xlabel('Count of Words', fontsize=12)
plt.show()

str1 = " " 
top_people = str1.join(birlestir1)

top_people = nlp(top_people)

label = [(X.text, X.label_) for X in top_people.ents]

df3 = pd.DataFrame(label, columns = ['Word','Entity'])

df3 = df3.where(df3['Entity'] == 'PERSON')

df4 = df3['Word'].value_counts()

print(df3)
print(df4)
print(label)

top_people = df4[:10,]
plt.figure(figsize=(10,5))
sns.barplot(top_people.values, top_people.index, alpha=0.8)
plt.title('Top People Mentioned')
plt.ylabel('Word from Tweet', fontsize=12)
plt.xlabel('Count of Words', fontsize=12)
plt.show()
print(top_people)

with open("twet.txt","r") as file:
    liste = file.read()

lines = list()
for line in liste:    
    words = line.split()
    for w in words: 
       lines.append(w)
from nltk.tokenize import sent_tokenize,word_tokenize
words=word_tokenize(liste)
print(words)

from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize

tokens_without_sw = [word for word in words if not word in stopwords.words('english')]
print(tokens_without_sw)

str1 = " " 
top_people = str1.join(tokens_without_sw)

top_people = nlp(top_people)

label = [(X.text, X.label_) for X in top_people.ents]

df5 = pd.DataFrame(label, columns = ['Word','Entity'])

df5 = df5.where(df5['Entity'] == 'PERSON') 

df6 = df5['Word'].value_counts()

top_people = df6[:20,]
plt.figure(figsize=(10,5))
sns.barplot(top_people.values, top_people.index, alpha=0.8)
plt.title('Top People Mentioned')
plt.ylabel('Word from Tweet', fontsize=12)
plt.xlabel('Count of Words', fontsize=12)
plt.show()
print(top_people)

str1 = " " 
top_org = str1.join(tokens_without_sw)

top_org = nlp(top_org)

label = [(X.text, X.label_) for X in top_org.ents]
print(label)

df7 = pd.DataFrame(label, columns = ['Word','Entity'])

df8 = df7.where(df7['Entity'] == 'ORG')
 
df8 = df8['Word'].value_counts()
print(df8)

top_org = df8[:20,]
plt.figure(figsize=(10,5))
sns.barplot(top_org.values, top_org.index, alpha=0.8)
plt.title('Top Organizations Mentioned')
plt.ylabel('Word from Tweet', fontsize=12)
plt.xlabel('Count of Words', fontsize=12)
plt.show()